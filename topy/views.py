from asyncore import read
from operator import inv
from django.http import HttpResponse
from unicodedata import name
from urllib import response
from django.shortcuts import render
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment
from django.http import JsonResponse, request, FileResponse
import os
import time


def index(request):
    if "GET" == request.method:
        return render(request, 'topy/index.html',{})
    else : 
        #start_work = time.time()
        #topy = core(request)
        file_path = "./topy.xlsx"
        response = HttpResponse(open(file_path, 'rb').read())
        #print(response)
        response['Content-Type'] = 'mimetype/submimetype'
        response['Content-Disposition'] = 'attachment; filename=topy.xlsx'
        #print('all time',time.time()-start_work)
        return response
def core(request):
    
    excel_file = request.FILES["excel_file"]
    wb = openpyxl.load_workbook(excel_file, read_only=True)
    
    #read_verible = time.time()
    transaction_sh = wb["transaction"]
    transactionarchiw_sh = wb["archiw"]
    inventory_sh = wb["inventory"]
    ean_sh = wb["sku"]
    topy_sh = wb["topy"]
    #print ("read verible",(time.time()-read_verible))
    
    #analise_verible = time.time()
    topy = get_sku_topy(topy_sh)

    # transaction_time = time.time()
    transaction_archiw = read_transaction(transactionarchiw_sh)
    transaction = read_transaction(transaction_sh)
    #print("transacktion time ", (time.time()-transaction_time))

    inventory = get_inventory(inventory_sh)
    ean = get_ean(ean_sh)
    name = get_names(topy_sh)
    #print ("analise verible",(time.time()-analise_verible))

    #print(name)
    #print(topy[0])
    #print(transaction[topy[1]],transaction_archiw[topy[1]],inventory[topy[1]],ean[topy[1]])
    #print(transaction[topy[1]],transaction_archiw[topy[1]])
    #topys = resoltb.active

    get_resolt = time.time()
    path = './topy.xlsx'
    try:
        os.remove(path)
        #print("Deleted")
    except:
        None

    resoltb = Workbook()
    topys = resoltb.create_sheet("Topy",0)
    row = 1
    for top in topy:
        #top = 45887233
        # провірка на існування списку з адресами 
        if top in transaction and top in transaction_archiw:
            #print ('ok',transaction_archiw[top])
            adreses = unical_adres(transaction[top],transaction_archiw[top])
        elif top not in transaction and top in transaction_archiw:
            adreses = transaction_archiw[top]
        elif top in transaction and top not in transaction_archiw:
            adreses = transaction[top]
        else :
            adreses = []
        if inventory.get(top)!=None:
           aq = adresqty(adreses,inventory[top])
        else:
            aq = adreses
       #створення рядка для запису і одній клітинці всі адресів
        adrqtytowrite = ''
        name_ean = ""+name[top]+'\n'
        for a in aq:
            adrqtytowrite+=(a+'\n')
        #print(top,aq,ean[top])
        #print('\n')
        # додавання до адресів ean
        for e in ean[top]:
            name_ean+=(str(e)+'\n')
        #print(name_ean)
        #topys.cell('A1').style.alignment.wrap_text = True
        topys[f'C{row}'].alignment = Alignment(wrapText=True)
        topys[f'B{row}'].alignment = Alignment(wrapText=True)
        topys[f'A{row}']=top
        topys[f'B{row}']=name_ean
        #topys[f'B{row}'].alignment = Alignment(wrapText=True)
        topys[f'c{row}']=adrqtytowrite
        row+=1
    resoltb.save('topy.xlsx')
    #print ("get resolt",(time.time()-get_resolt))
    #print("all time",(time.time()-start_read))
    #print(type(resoltb))
    return (resoltb)    



def read_transaction(transaction_sh):
   
    maxindex = 'D'+str(transaction_sh.max_row)
    table = transaction_sh['B1':maxindex]
    transaction_dict = {}
    for s,fromloc,toloc in table:
        # if s.value not in transaction_dict  :
        #     transaction_dict[s.value] = []
        #     if filter(fromloc.value):
        #         transaction_dict[s.value].append(fromloc.value)
        #     if filter(toloc.value):
        #         transaction_dict[s.value].append(toloc.value)
        # else :
        #     if  fromloc.value not in transaction_dict[s.value] and filter(fromloc.value):
        #         transaction_dict[s.value].append(fromloc.value)
        #     if  toloc.value not in transaction_dict[s.value] and filter(toloc.value):
        #         transaction_dict[s.value].append(toloc.value)
        
        if s.value not in transaction_dict:
            transaction_dict[s.value] = []
            transaction_dict[s.value].append(str(fromloc.value))
            transaction_dict[s.value].append(str(toloc.value))
        else:
            transaction_dict[s.value].append(str(fromloc.value))
            transaction_dict[s.value].append(str(toloc.value))

    for sku in transaction_dict:
        betwin_list=[]
        for adres in list(set(transaction_dict[sku])):
            if filter(adres):
                betwin_list.append(adres)
        transaction_dict[sku]=betwin_list

    return (transaction_dict)

def get_sku_topy(topy_sh):
    #sku_row = topy_sh['A']
    sku_list=[]
    for sku in range(1,topy_sh.max_row+1):
        #print(topy_sh[sku][0].value)
        row_value = topy_sh[sku][0].value
        #sku_list.append(sku.value) if type(sku.value)==int and sku.value not in sku_list else None
        sku_list.append(row_value) if type(row_value)==int and row_value not in sku_list else None
    
    return sku_list

def get_inventory(inventory_sh):
    # sku_row = inventory_sh["F"]
    # location_row = inventory_sh["G"]
    # quty_row = inventory_sh["H"]
    #print("inventoru###########")
    # sku_row = []
    # location_row = []
    # quty_row = []

    # for row in range(1, inventory_sh.max_row+1):
    #     sku = inventory_sh[row][5]
    #     location = inventory_sh[row][6]
    #     quty = inventory_sh[row][7]

    #     sku_row.append(sku)
    #     location_row.append(location)
    #     quty_row.append(quty)
        #print(sku.value,'-',location.value,'-',quty.value)

    maxindex = 'H'+str(inventory_sh.max_row)
    table = inventory_sh['F1':maxindex]

    # створюю не фільтрований  словник inventory_dict = {sku:[[adres,quty],[adres1,quty]]...sku_n:[[adres_n,quty],[adres_n+1,quty]]}
    inventory_dict_diorty = {}
    inventory_dict_clin = {}
    for s_row, l_row, q_row in table:
    #for s_row, l_row, q_row in zip(sku_row,location_row,quty_row):
        if s_row.value not in inventory_dict_diorty:
            if s_row.value != None:
                inventory_dict_diorty[s_row.value]=[[l_row.value,q_row.value]]
        else:
            if s_row.value != None:
                inventory_dict_diorty[s_row.value].append([l_row.value,q_row.value])
    # slq - sku + location + quanuty 
    # сумую співпвдіння по адресах ,1 адркс = загальна к-ть
    #print(inventory_dict_diorty)
    for slq in inventory_dict_diorty:
        betwin_dickt = {}
        for adres_qty in inventory_dict_diorty[slq]:
            #print(adres_qty[0])
            if adres_qty[0] not in betwin_dickt:
                betwin_dickt[adres_qty[0]]=adres_qty[1]
            else:
                # print('!!!!!!!!!!',adres_qty[0])
                # print("@@@@@@@@@@-",adres_qty[1])
                # print(adres_qty)
                betwin_dickt[adres_qty[0]]+=adres_qty[1]
        clean_list = []
        for adres in betwin_dickt:
            if filter(adres):
                clean_list.append([adres,betwin_dickt[adres]])
        inventory_dict_clin[slq]=clean_list
    # for i in inventory_dict_clin:
    #     print(i,'--',inventory_dict_clin[i])
    return (inventory_dict_clin)
   
def get_ean(ean_sh):
    # sku_row = ean_sh["A"]
    # ean_row = ean_sh["D"]

    sku_row = []
    ean_row = []

    for row in range(1,ean_sh.max_row+1):
       sku = ean_sh[row][0]
       ean = ean_sh[row][4]
       sku_row.append(sku)
       ean_row.append(ean)

    ean_list = []
    suplier_dict = {}
    for s_row, e_row in zip(sku_row,ean_row):
        if e_row.value not in ean_list:
            ean_list.append(e_row.value)
            if s_row.value in suplier_dict :
                suplier_dict[s_row.value].append(e_row.value)
            else:
               suplier_dict[s_row.value]= [e_row.value]
    return(suplier_dict)
# обєднання адресів з врхіву і трансакшенів
def unical_adres(transaction , archiw):
    adres_list=archiw
    for adres in transaction:
        if adres not in archiw:
            adres_list.append(adres)
    return (adres_list)
# стврення списку  адресів і кілткості товару 
def adresqty(adreses, inventory):
    final_list = []
    not_add_list = []
    for aq in inventory:
        #print(aq)
        #print(aq[0]," - ",str(aq[1]))
        final_list.append(aq[0]+" - "+str(aq[1]))
        not_add_list.append(aq[0])
    for adres in adreses:
        if adres not in not_add_list:
            final_list.append(adres) 
    return (final_list)

def get_names(topy_sh):
    # name_row = topy_sh['B']
    # sku_row = topy_sh['A']

    # name_row = []
    # sku_row = []

    # for row in range(1,topy_sh.max_row+1):
    #     name = topy_sh[row][1]
    #     sku = topy_sh[row][0]

    #     name_row.append(name)
    #     sku_row.append(sku)


    name_dict = {}

    maxindex = 'B'+str(topy_sh.max_row)
    table = topy_sh['A1':maxindex]

    for sku, name in table:
    #or sku, name in zip(sku_row,name_row):
        if type(sku.value)==int:
            name_dict[sku.value]=name.value
            #print(name.value)
    return name_dict

def filter(adres):
    #print(adres)
    adres = str(adres)
    adres = adres.replace('-','')
    if adres=='' or adres == None or len(adres)<=4:
        return False
    # куветки   K*******  K1204503  // вяти      W**** W0123 // карнизи   D**** D1330 
    if adres[0]=='K' or adres[0]=='W' or adres[0]=='D':
        if adres[1:].isdigit():
            return True
        else:
            return False

    elif (adres[0:1]=="WK" or adres[0:1]=="WS") and adres[2:].isdigit():
        return True
            
    # регали    **R****L** 01R2104A10 01R1027A15 01R0517H1
    elif adres!="PARKING" and adres[2]=='R' and adres[7].isalpha():
        if adres[0:2].isdigit() and adres[3:7].isdigit():
            return True
        else:
            return False
    # маси      **M*** 01M242   01M525-20 01M316-03
    elif adres[2]=='M':
        if adres[3:].isdigit():
            return True
        else:
            return False
    elif adres[:8]=="SZYB-ZAM" :
        return True
    elif adres=='INRACK80':
        return True
    else:
        return False
        